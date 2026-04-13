import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_submissions_xlsx(form, submissions_qs):
    """Generate XLSX with dynamic columns matching form fields + metadata."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Submissions'

    # Collect active fields in order to define columns
    form_fields = list(
        form.fields.filter(is_active=True)
        .exclude(field_type__in=['section_heading', 'description_block'])
        .order_by('sort_order', 'created_at')
    )

    # Headers: metadata + dynamic field columns
    meta_headers = ['#', 'Submitted At', 'IP Address', 'City', 'Country', 'Device', 'Browser', 'OS', 'Status']
    field_headers = [f.label for f in form_fields]
    all_headers = meta_headers + field_headers

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    field_id_list = [str(f.id) for f in form_fields]

    for row_idx, sub in enumerate(submissions_qs, 2):
        # Build answer lookup by question_id
        answer_map = {}
        for ans in sub.answers.all():
            if ans.question_id:
                answer_map[str(ans.question_id)] = ans.answer_value or (
                    str(ans.answer_json) if ans.answer_json else ''
                )

        meta_values = [
            row_idx - 1,
            sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if sub.submitted_at else '',
            sub.ip_address or '',
            sub.city,
            sub.country,
            sub.device_type,
            sub.browser_name,
            sub.os_name,
            sub.get_status_display(),
        ]
        field_values = [answer_map.get(fid, '') for fid in field_id_list]
        row_data = meta_values + field_values

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-width
    for col_idx, header in enumerate(all_headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
