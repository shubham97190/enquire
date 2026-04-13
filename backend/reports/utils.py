import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HEADERS = [
    'Name',
    'Email',
    'Phone',
    'Category',
    'Description',
    'OTP Verified',
    'Created At',
    'Status',
    'Admin Notes',
    'Download Count',
]


def generate_enquiry_xlsx(enquiries):
    """Generate an XLSX workbook from enquiry queryset."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Enquiries'

    # Header styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, enquiry in enumerate(enquiries, 2):
        row_data = [
            enquiry.name,
            enquiry.email,
            enquiry.phone,
            enquiry.category.name if enquiry.category else '',
            enquiry.description,
            'Yes' if enquiry.otp_verified else 'No',
            enquiry.created_at.strftime('%Y-%m-%d %H:%M:%S') if enquiry.created_at else '',
            enquiry.get_status_display(),
            enquiry.admin_notes,
            enquiry.download_count,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, header in enumerate(HEADERS, 1):
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
